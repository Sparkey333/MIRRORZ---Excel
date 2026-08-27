import re, sys, subprocess, shutil, os
from pathlib import Path
from openpyxl import Workbook

cases = [l.rstrip("\n") for l in open(sys.argv[1]) if l.strip() and not l.startswith("#")]
wb = Workbook()
d = wb.active
d.title = "D"
for i, v in enumerate([4, 5, 8, 7, 11, 4, 3, 9], start=1):
    d.cell(row=i, column=1, value=v)
for i, v in enumerate([2, 4, 6, 8, 10, 12, 14, 16], start=1):
    d.cell(row=i, column=2, value=v)
for i, v in enumerate(["a", "b", "a", "c"], start=1):
    d.cell(row=i, column=3, value=v)
for i, v in enumerate([1, 2, 3, 4, 5, 6, 7, 8], start=1):
    d.cell(row=i, column=4, value=v)
d.cell(row=1, column=5, value=10)
d.cell(row=2, column=5, value="x")
d.cell(row=3, column=5, value=True)
d.cell(row=4, column=5, value="")
d.cell(row=5, column=5, value=20)
# G: y values for regression, H: x values
for i, v in enumerate([2.1, 4.3, 5.9, 8.4, 10.1], start=1):
    d.cell(row=i, column=7, value=v)
for i, v in enumerate([1, 2, 3, 4, 5], start=1):
    d.cell(row=i, column=8, value=v)
# I,J: second x variable and a second sample
for i, v in enumerate([3, 1, 4, 1, 5], start=1):
    d.cell(row=i, column=9, value=v)
for i, v in enumerate([6, 9, 4, 8, 7, 5], start=1):
    d.cell(row=i, column=10, value=v)

reg = open("packages/formula/src/registry.ts").read()
body = reg.split("FUTURE_FUNCTIONS: ReadonlySet<string> = new Set([")[1].split("]);")[0]
FUTURE = set(re.findall(r"'([A-Z][A-Z0-9._]*)'", body))

def prefix(f):
    return re.sub(r"\b([A-Z][A-Z0-9._]*)\s*\(",
                  lambda m: ("_xlfn." + m.group(1) if (m.group(1) in FUTURE or "." in m.group(1)) else m.group(1)) + "(", f)

c = wb.create_sheet("C")
for i, f in enumerate(cases, start=1):
    c.cell(row=i, column=1, value=f"case{i}")
    c.cell(row=i, column=2, value="=" + prefix(f))
out = Path("scratch-tmp/probe.xlsx")
wb.save(out)

env = dict(os.environ, HOME=os.environ.get("HOME") or "/root")
subprocess.run(["soffice", "--headless", "--norestore",
                "-env:UserInstallation=file:///tmp/mirrorz-lo-profile",
                "--convert-to", "xlsx:Calc MS Excel 2007 XML",
                "--outdir", "scratch-tmp/_out", str(out)],
               check=True, capture_output=True, timeout=300, env=env)
shutil.move("scratch-tmp/_out/probe.xlsx", "scratch-tmp/probe.calc.xlsx")
print("ok", len(cases))
