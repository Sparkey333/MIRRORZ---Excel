#!/usr/bin/env python3
"""Generate reference spreadsheet fixtures for MIRRORZ round-trip and oracle tests.

Two-stage pipeline:
  1. openpyxl authors feature-rich .xlsx workbooks with KNOWN expected content.
  2. LibreOffice headless converts each to .xls / .ods / .csv / .html so the same
     logical content exists in every format we claim to read.

LibreOffice is used only as an offline build/test tool - it is never linked into
or shipped with MIRRORZ, so its LGPL/MPL terms do not affect our Apache-2.0 code.
"""
import json
import os
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, IconSetRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "fixtures" / "generated"


# --- OOXML "future function" prefixing -------------------------------------
# Functions introduced after Excel 2007 are stored in the XML with an _xlfn.
# prefix (and _xlfn._xlws. for a few worksheet-only ones). Excel strips the
# prefix for display; a reader that does not strip it reports #NAME?. We apply
# the prefix here so fixtures are byte-honest about what real files contain,
# which also exercises our own reader's prefix handling.
XLFN = {
    "IFS", "XOR", "TEXTJOIN", "CONCAT", "SWITCH", "MAXIFS", "MINIFS", "IFNA",
    "STDEV.S", "STDEV.P", "VAR.S", "VAR.P", "PERCENTILE.INC", "PERCENTILE.EXC",
    "QUARTILE.INC", "QUARTILE.EXC", "RANK.EQ", "RANK.AVG", "MODE.SNGL", "MODE.MULT",
    "NORM.DIST", "NORM.INV", "NORM.S.DIST", "NORM.S.INV", "T.TEST", "F.TEST",
    "CHISQ.TEST", "COVARIANCE.P", "COVARIANCE.S", "BINOM.DIST", "EXPON.DIST",
    "CEILING.MATH", "FLOOR.MATH", "CEILING.PRECISE", "FLOOR.PRECISE",
    "XLOOKUP", "XMATCH", "LET", "LAMBDA", "BYROW", "BYCOL", "MAP", "REDUCE",
    "SCAN", "MAKEARRAY", "ISOMITTED", "UNIQUE", "SORT", "SORTBY", "SEQUENCE",
    "RANDARRAY", "ARRAYTOTEXT", "VALUETOTEXT", "TEXTSPLIT", "TEXTBEFORE",
    "TEXTAFTER", "VSTACK", "HSTACK", "TOCOL", "TOROW", "CHOOSECOLS",
    "CHOOSEROWS", "WRAPROWS", "WRAPCOLS", "EXPAND", "TAKE", "DROP",
    "REGEXTEST", "REGEXEXTRACT", "REGEXREPLACE", "GROUPBY", "PIVOTBY",
    "PERCENTOF", "TRIMRANGE", "DAYS", "ISOWEEKNUM",     "NUMBERVALUE", "UNICHAR", "UNICODE", "BASE", "DECIMAL", "COMBINA", "PERMUTATIONA",
    "SEC", "CSC", "COT", "ACOT", "SECH", "CSCH", "COTH", "ARABIC", "BITAND", "BITOR",
    "BITXOR", "BITLSHIFT", "BITRSHIFT", "IMTAN", "IMCOSH", "IMSINH", "IMSEC", "IMCSC",
    "PDURATION", "RRI", "FORMULATEXT", "SHEET", "SHEETS", 
}
# These live under the worksheet namespace as well.
XLWS = {"FILTER", "ANCHORARRAY", "SINGLE"}

_FN_RE = __import__("re").compile(r"(?<![A-Za-z0-9_.$!])([A-Z][A-Z0-9._]*)\s*\(")


def xlfn(formula: str) -> str:
    """Rewrite bare future-function names into their stored _xlfn. form."""
    def sub(m):
        name = m.group(1)
        if name in XLWS:
            return f"_xlfn._xlws.{name}("
        if name in XLFN:
            return f"_xlfn.{name}("
        return m.group(0)
    return _FN_RE.sub(sub, formula)


def basic_types(path: Path) -> dict:
    """Every primitive cell type plus the classic Excel numeric edge cases."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Types"
    rows = [
        ("label", "value", "note"),
        ("integer", 42, "plain int"),
        ("negative", -17, "negative int"),
        ("zero", 0, "zero"),
        ("float", 3.14159265358979, "pi to 15 sig digits"),
        ("tiny", 1e-300, "denormal-ish"),
        ("huge", 1e300, "near float max"),
        ("string", "hello world", "shared string"),
        ("unicode", "éàü 你好 \U0001f600", "accents, CJK, emoji"),
        ("quote", 'he said "hi"', "embedded quotes"),
        ("bool_true", True, "boolean"),
        ("bool_false", False, "boolean"),
        ("empty", None, "blank cell"),
        ("leading_zero", "007", "must stay text"),
        ("looks_like_date", "1-2", "must stay text, not Jan 2"),
        ("gene_name", "SEPT1", "must stay text, not a date"),
        ("long_number", "1234567890123456789", "beyond 15 sig digits, as text"),
    ]
    for r in rows:
        ws.append(list(r))
    ws["B18"] = 0.1
    ws["B19"] = 0.2
    ws["B20"] = "=B18+B19"
    ws["A18"], ws["A19"], ws["A20"] = "point one", "point two", "sum (0.3 exactly in Excel)"
    ws["A21"], ws["B21"] = "date serial 1", "=DATE(1900,1,1)"
    ws["A22"], ws["B22"] = "leap bug serial 60", "=DATE(1900,2,28)+1"
    wb.save(path)
    return {"sheets": ["Types"], "focus": "primitive types + numeric/date edge cases"}


def formulas(path: Path) -> dict:
    """A broad formula surface. LibreOffice will cache the correct values for us."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["n", "name", "dept", "salary", "hired"])
    data = [
        (1, "Ada", "Eng", 165000, "2019-03-01"),
        (2, "Grace", "Eng", 172000, "2018-07-15"),
        (3, "Katherine", "Science", 158000, "2020-01-20"),
        (4, "Dorothy", "Science", 149000, "2021-11-02"),
        (5, "Margaret", "Eng", 181000, "2017-05-30"),
        (6, "Annie", "Ops", 121000, "2022-09-12"),
        (7, "Henrietta", "Ops", 118000, "2023-02-28"),
        (8, "Mary", "Science", 143000, "2020-06-08"),
    ]
    for row in data:
        ws.append(list(row))

    fx = wb.create_sheet("Formulas")
    cases = [
        ("SUM", "=SUM(Data!D2:D9)"),
        ("AVERAGE", "=AVERAGE(Data!D2:D9)"),
        ("MIN", "=MIN(Data!D2:D9)"),
        ("MAX", "=MAX(Data!D2:D9)"),
        ("COUNT", "=COUNT(Data!D2:D9)"),
        ("COUNTA", "=COUNTA(Data!A1:E9)"),
        ("COUNTBLANK", "=COUNTBLANK(Data!A1:E20)"),
        ("MEDIAN", "=MEDIAN(Data!D2:D9)"),
        ("STDEV.S", "=STDEV.S(Data!D2:D9)"),
        ("VAR.S", "=VAR.S(Data!D2:D9)"),
        ("SUMIF", '=SUMIF(Data!C2:C9,"Eng",Data!D2:D9)'),
        ("SUMIFS", '=SUMIFS(Data!D2:D9,Data!C2:C9,"Eng",Data!D2:D9,">170000")'),
        ("COUNTIF", '=COUNTIF(Data!C2:C9,"Science")'),
        ("COUNTIFS", '=COUNTIFS(Data!C2:C9,"Ops",Data!D2:D9,"<120000")'),
        ("AVERAGEIF", '=AVERAGEIF(Data!C2:C9,"Eng",Data!D2:D9)'),
        ("IF", '=IF(Data!D2>160000,"high","low")'),
        ("IFS", '=IFS(Data!D2>180000,"top",Data!D2>150000,"mid",TRUE,"base")'),
        ("AND", "=AND(1=1,2=2)"),
        ("OR", "=OR(1=2,2=2)"),
        ("NOT", "=NOT(1=2)"),
        ("XOR", "=XOR(TRUE,FALSE)"),
        ("IFERROR", '=IFERROR(1/0,"caught")'),
        ("IFNA", '=IFNA(NA(),"was na")'),
        ("VLOOKUP", '=VLOOKUP(3,Data!A2:D9,2,FALSE)'),
        ("HLOOKUP", '=HLOOKUP("dept",Data!A1:E9,4,FALSE)'),
        ("INDEX_MATCH", '=INDEX(Data!D2:D9,MATCH("Grace",Data!B2:B9,0))'),
        ("XLOOKUP", '=XLOOKUP("Annie",Data!B2:B9,Data!D2:D9,"none")'),
        ("XMATCH", '=XMATCH("Mary",Data!B2:B9,0)'),
        ("OFFSET", "=OFFSET(Data!A1,2,3)"),
        ("INDIRECT", '=INDIRECT("Data!D2")'),
        ("CHOOSE", '=CHOOSE(2,"a","b","c")'),
        ("ROW", "=ROW(Data!A5)"),
        ("COLUMN", "=COLUMN(Data!D1)"),
        ("ROWS", "=ROWS(Data!A1:A9)"),
        ("COLUMNS", "=COLUMNS(Data!A1:E1)"),
        ("CONCAT", '=CONCAT("a","b","c")'),
        ("TEXTJOIN", '=TEXTJOIN("-",TRUE,"x","y","z")'),
        ("LEFT", '=LEFT("spreadsheet",6)'),
        ("RIGHT", '=RIGHT("spreadsheet",5)'),
        ("MID", '=MID("spreadsheet",7,5)'),
        ("LEN", '=LEN("spreadsheet")'),
        ("UPPER", '=UPPER("abc")'),
        ("LOWER", '=LOWER("ABC")'),
        ("PROPER", '=PROPER("hello world")'),
        ("TRIM", '=TRIM("  pad  ")'),
        ("SUBSTITUTE", '=SUBSTITUTE("a-b-c","-","+")'),
        ("REPLACE", '=REPLACE("abcdef",2,3,"XY")'),
        ("FIND", '=FIND("read","spreadsheet")'),
        ("SEARCH", '=SEARCH("READ","spreadsheet")'),
        ("TEXT", '=TEXT(1234.567,"#,##0.00")'),
        ("VALUE", '=VALUE("123.45")'),
        ("REPT", '=REPT("ab",3)'),
        ("EXACT", '=EXACT("a","A")'),
        ("ROUND", "=ROUND(3.14159,2)"),
        ("ROUNDUP", "=ROUNDUP(3.14159,2)"),
        ("ROUNDDOWN", "=ROUNDDOWN(3.14159,2)"),
        ("MROUND", "=MROUND(17,5)"),
        ("CEILING.MATH", "=CEILING.MATH(4.2)"),
        ("FLOOR.MATH", "=FLOOR.MATH(4.8)"),
        ("INT", "=INT(-3.5)"),
        ("TRUNC", "=TRUNC(-3.5)"),
        ("ABS", "=ABS(-7)"),
        ("SIGN", "=SIGN(-7)"),
        ("MOD", "=MOD(-7,3)"),
        ("QUOTIENT", "=QUOTIENT(7,2)"),
        ("POWER", "=POWER(2,10)"),
        ("SQRT", "=SQRT(144)"),
        ("EXP", "=EXP(1)"),
        ("LN", "=LN(EXP(2))"),
        ("LOG10", "=LOG10(1000)"),
        ("LOG", "=LOG(8,2)"),
        ("PI", "=PI()"),
        ("SIN", "=SIN(PI()/2)"),
        ("COS", "=COS(0)"),
        ("TAN", "=TAN(0)"),
        ("ATAN2", "=ATAN2(1,1)"),
        ("DEGREES", "=DEGREES(PI())"),
        ("RADIANS", "=RADIANS(180)"),
        ("SUMPRODUCT", "=SUMPRODUCT(Data!A2:A9,Data!D2:D9)"),
        ("SUMSQ", "=SUMSQ(3,4)"),
        ("PRODUCT", "=PRODUCT(2,3,4)"),
        ("GCD", "=GCD(24,36)"),
        ("LCM", "=LCM(4,6)"),
        ("FACT", "=FACT(6)"),
        ("COMBIN", "=COMBIN(10,3)"),
        ("DATE", "=DATE(2024,2,29)"),
        ("YEAR", "=YEAR(DATE(2024,2,29))"),
        ("MONTH", "=MONTH(DATE(2024,2,29))"),
        ("DAY", "=DAY(DATE(2024,2,29))"),
        ("EOMONTH", "=EOMONTH(DATE(2024,1,15),0)"),
        ("EDATE", "=EDATE(DATE(2024,1,31),1)"),
        ("WEEKDAY", "=WEEKDAY(DATE(2024,3,1))"),
        ("WEEKNUM", "=WEEKNUM(DATE(2024,3,1))"),
        ("DATEDIF", '=DATEDIF(DATE(2020,1,1),DATE(2024,3,1),"m")'),
        ("DAYS", "=DAYS(DATE(2024,3,1),DATE(2024,1,1))"),
        ("NETWORKDAYS", "=NETWORKDAYS(DATE(2024,1,1),DATE(2024,1,31))"),
        ("WORKDAY", "=WORKDAY(DATE(2024,1,1),10)"),
        ("TIME", "=TIME(13,45,30)"),
        ("HOUR", "=HOUR(TIME(13,45,30))"),
        ("MINUTE", "=MINUTE(TIME(13,45,30))"),
        ("SECOND", "=SECOND(TIME(13,45,30))"),
        ("PMT", "=PMT(0.05/12,360,-300000)"),
        ("FV", "=FV(0.05/12,360,-1000)"),
        ("PV", "=PV(0.05/12,360,-1000)"),
        ("NPER", "=NPER(0.05/12,-1000,200000)"),
        ("RATE", "=RATE(360,-1500,250000)"),
        ("NPV", "=NPV(0.1,-100,50,60,70)"),
        ("IRR", "=IRR({-100,50,60,70})"),
        ("SLN", "=SLN(10000,1000,5)"),
        ("ISBLANK", "=ISBLANK(Data!Z1)"),
        ("ISNUMBER", "=ISNUMBER(1)"),
        ("ISTEXT", '=ISTEXT("a")'),
        ("ISERROR", "=ISERROR(1/0)"),
        ("ISERR", "=ISERR(NA())"),
        ("ISNA", "=ISNA(NA())"),
        ("ISLOGICAL", "=ISLOGICAL(TRUE)"),
        ("ISEVEN", "=ISEVEN(4)"),
        ("ISODD", "=ISODD(4)"),
        ("N", "=N(TRUE)"),
        ("T", '=T("txt")'),
        ("TYPE", "=TYPE(1)"),
        ("ERR_DIV0", "=1/0"),
        ("ERR_VALUE", '="a"+1'),
        ("ERR_NAME", "=NOTAFUNCTION(1)"),
        ("ERR_REF", "=SUM(Data!A1:A2)/0"),
        ("ARRAY_SUM", "=SUM({1,2,3;4,5,6})"),
        ("NESTED", "=IF(SUM(Data!D2:D9)>1000000,ROUND(AVERAGE(Data!D2:D9),0),0)"),
        ("PCT_RANK", "=PERCENTILE.INC(Data!D2:D9,0.5)"),
        ("RANK", "=RANK.EQ(Data!D2,Data!D2:D9)"),
        ("LARGE", "=LARGE(Data!D2:D9,2)"),
        ("SMALL", "=SMALL(Data!D2:D9,2)"),
        ("ABS_REF", "=Data!$D$2+1"),
        ("RANGE_OP", "=SUM(Data!D2:D5 Data!D4:D9)"),
        ("UNION_OP", "=SUM(Data!D2:D3,Data!D8:D9)"),
        ("PERCENT_LIT", "=50%"),
        ("EXPONENT", "=2^10"),
        ("CONCAT_OP", '="a"&"b"'),
        ("COMPARE", "=1<2"),
        ("UNARY_NEG", "=-(3+4)"),
    ]
    fx.append(["case", "formula", "result"])
    for i, (name, formula) in enumerate(cases, start=2):
        fx.cell(row=i, column=1, value=name)
        fx.cell(row=i, column=2, value=formula.replace("=", "'=", 1))
        fx.cell(row=i, column=3, value=xlfn(formula))
    wb.save(path)
    return {"sheets": ["Data", "Formulas"], "cases": len(cases),
            "focus": "formula breadth; col C holds live formulas, col B the same text for reference"}


def styling(path: Path) -> dict:
    wb = Workbook()
    ws = wb.active
    ws.title = "Styles"
    ws["A1"] = "bold"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = "italic"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = "underline"
    ws["A3"].font = Font(underline="single")
    ws["A4"] = "strike"
    ws["A4"].font = Font(strike=True)
    ws["A5"] = "big red"
    ws["A5"].font = Font(size=20, color="FFCC0000", name="Arial")
    ws["B1"] = "yellow fill"
    ws["B1"].fill = PatternFill("solid", fgColor="FFFFFF00")
    ws["B2"] = "bordered"
    thin = Side(style="thin", color="FF000000")
    thick = Side(style="thick", color="FF0000FF")
    ws["B2"].border = Border(left=thin, right=thin, top=thick, bottom=thick)
    ws["B3"] = "centered"
    ws["B3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B4"] = "wrapped text that is quite long and should wrap onto several lines"
    ws["B4"].alignment = Alignment(wrap_text=True)
    ws["B5"] = "rotated"
    ws["B5"].alignment = Alignment(textRotation=45)
    ws["B6"] = "indented"
    ws["B6"].alignment = Alignment(indent=3)

    fmts = [
        ("general", 1234.5678, "General"),
        ("2dp", 1234.5678, "0.00"),
        ("thousands", 1234567.891, "#,##0.00"),
        ("currency", 1234.5, '"$"#,##0.00'),
        ("accounting", -1234.5, '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'),
        ("percent", 0.4567, "0.00%"),
        ("scientific", 12345.678, "0.00E+00"),
        ("fraction", 0.75, "# ?/?"),
        ("date", 45000, "yyyy-mm-dd"),
        ("date_long", 45000, "dddd, mmmm d, yyyy"),
        ("time", 0.5678, "hh:mm:ss"),
        ("datetime", 45000.5678, "yyyy-mm-dd hh:mm"),
        ("elapsed", 1.5, "[h]:mm:ss"),
        ("negred", -42, "0.00;[Red]-0.00"),
        ("four_section", 0, '0.00;[Red]-0.00;"zero";@'),
        ("text_fmt", "raw", "@"),
        ("conditional", 150, "[>100]\"big\";[<=100]\"small\""),
    ]
    ws["D1"], ws["E1"], ws["F1"] = "name", "value", "numFmt"
    for i, (name, value, fmt) in enumerate(fmts, start=2):
        ws.cell(row=i, column=4, value=name)
        c = ws.cell(row=i, column=5, value=value)
        c.number_format = fmt
        ws.cell(row=i, column=6, value=fmt)

    ws.merge_cells("H1:J2")
    ws["H1"] = "merged H1:J2"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.row_dimensions[4].height = 48
    ws.freeze_panes = "C3"
    ws.column_dimensions["G"].hidden = True
    ws.row_dimensions[10].hidden = True
    wb.save(path)
    return {"sheets": ["Styles"], "numberFormats": len(fmts),
            "focus": "fonts, fills, borders, alignment, number-format mini-language, merges, panes"}


def features(path: Path) -> dict:
    wb = Workbook()
    ws = wb.active
    ws.title = "Features"
    ws.append(["region", "qty", "score"])
    for i, (region, qty, score) in enumerate(
        [("North", 12, 0.91), ("South", 45, 0.42), ("East", 78, 0.67),
         ("West", 33, 0.15), ("North", 91, 0.88), ("South", 5, 0.30)], start=2):
        ws.append([region, qty, score])

    ws.conditional_formatting.add(
        "B2:B7", CellIsRule(operator="greaterThan", formula=["50"],
                            fill=PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")))
    ws.conditional_formatting.add("C2:C7", ColorScaleRule(
        start_type="min", start_color="FFF8696B", mid_type="percentile", mid_value=50,
        mid_color="FFFFEB84", end_type="max", end_color="FF63BE7B"))
    ws.conditional_formatting.add("B2:B7", DataBarRule(
        start_type="min", end_type="max", color="FF638EC6", showValue=True))
    ws.conditional_formatting.add("C2:C7", IconSetRule("3TrafficLights1", "percent", [0, 33, 67]))

    dv = DataValidation(type="list", formula1='"North,South,East,West"', allow_blank=True)
    dv.error = "Pick a listed region"
    dv.errorTitle = "Invalid region"
    dv.prompt = "Choose a region"
    ws.add_data_validation(dv)
    dv.add("A2:A20")

    dv2 = DataValidation(type="whole", operator="between", formula1="0", formula2="100")
    ws.add_data_validation(dv2)
    dv2.add("B2:B20")

    tbl = Table(displayName="SalesTable", ref="A1:C7")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tbl)

    ws.auto_filter.ref = "A1:C7"
    ws["E1"] = "https://example.com"
    ws["E1"].hyperlink = "https://example.com"
    wb.defined_names.add(__import__("openpyxl").workbook.defined_name.DefinedName(
        "TotalQty", attr_text="Features!$B$2:$B$7"))
    ws["E3"] = "=SUM(TotalQty)"

    charts = wb.create_sheet("Charts")
    for r, row in enumerate([["cat", "s1", "s2"], ["a", 3, 7], ["b", 5, 2], ["c", 9, 4], ["d", 1, 8]], start=1):
        for c, v in enumerate(row, start=1):
            charts.cell(row=r, column=c, value=v)
    data = Reference(charts, min_col=2, min_row=1, max_col=3, max_row=5)
    cats = Reference(charts, min_col=1, min_row=2, max_row=5)
    bar = BarChart(); bar.title = "Bar"; bar.add_data(data, titles_from_data=True); bar.set_categories(cats)
    charts.add_chart(bar, "E2")
    line = LineChart(); line.title = "Line"; line.add_data(data, titles_from_data=True); line.set_categories(cats)
    charts.add_chart(line, "E20")
    pie = PieChart(); pie.title = "Pie"
    pie.add_data(Reference(charts, min_col=2, min_row=1, max_row=5), titles_from_data=True)
    pie.set_categories(cats)
    charts.add_chart(pie, "N2")

    notes = wb.create_sheet("Comments")
    notes["A1"] = "cell with comment"
    from openpyxl.comments import Comment
    notes["A1"].comment = Comment("This is a comment body", "MIRRORZ")
    notes.sheet_properties.tabColor = "FF00B050"

    hidden = wb.create_sheet("HiddenSheet")
    hidden["A1"] = "you should not see this by default"
    hidden.sheet_state = "hidden"
    wb.save(path)
    return {"sheets": ["Features", "Charts", "Comments", "HiddenSheet"],
            "focus": "conditional formatting, data validation, tables, autofilter, hyperlinks, defined names, charts, comments, hidden sheets, tab colour"}


def large(path: Path, rows: int = 50_000, cols: int = 20) -> dict:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Big")
    ws.append([f"col{c}" for c in range(1, cols + 1)])
    for r in range(rows):
        ws.append([r * c if c % 3 else f"s{r}-{c}" for c in range(1, cols + 1)])
    wb.save(path)
    return {"sheets": ["Big"], "rows": rows + 1, "cols": cols, "focus": "parser throughput and memory"}


def edge_cases(path: Path) -> dict:
    wb = Workbook()
    ws = wb.active
    ws.title = "Edge'\"Case"
    ws["A1"] = "sheet name has a quote and a double-quote"
    ws["A2"] = "=1"
    sparse = wb.create_sheet("Sparse")
    sparse["A1"] = "top-left"
    sparse["XFD1"] = "last column (16384)"
    sparse["A1048576"] = "last row"
    sparse["XFD1048576"] = "bottom-right corner"
    inline = wb.create_sheet("Wide")
    for c in range(1, 300):
        inline.cell(row=1, column=c, value=get_column_letter(c))
    ws2 = wb.create_sheet("Refs")
    ws2["A1"] = "=SUM('Edge''\"Case'!A2:A2)"
    ws2["A2"] = "=Sparse!A1"
    wb.save(path)
    return {"sheets": ["Edge'\"Case", "Sparse", "Wide", "Refs"],
            "focus": "quoted sheet names, extreme sparse addresses, wide sheets, cross-sheet refs"}


LO_PROFILE = "file:///tmp/mirrorz-lo-profile"


def _soffice(args: list, timeout: int = 240) -> None:
    env = dict(os.environ, HOME=os.environ.get("HOME") or "/root")
    subprocess.run(
        ["soffice", "--headless", "--norestore", f"-env:UserInstallation={LO_PROFILE}", *args],
        check=True, capture_output=True, timeout=timeout, env=env,
    )


CONVERSIONS = [
    ("xls", "MS Excel 97"),
    ("ods", "calc8"),
    ("csv", "Text - txt - csv (StarCalc)"),
]


def convert(src: Path, outdir: Path) -> list:
    made = []
    if not shutil.which("soffice"):
        return made
    for ext, filt in CONVERSIONS:
        try:
            _soffice(["--convert-to", f"{ext}:{filt}", "--outdir", str(outdir), str(src)])
            # soffice exits 0 even when the store step fails, so confirm the file exists.
            if (outdir / f"{src.stem}.{ext}").exists():
                made.append(f"{src.stem}.{ext}")
            else:
                print(f"  ! {src.name} -> {ext} produced no output", file=sys.stderr)
        except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as exc:
            print(f"  ! {src.name} -> {ext} failed: {exc}", file=sys.stderr)
    return made


def recalc(src: Path, outdir: Path) -> str | None:
    """Round-trip through LibreOffice so formula cells carry authoritative cached values.

    openpyxl writes formulas with NO cached <v>. LibreOffice recalculates on load and
    writes the values back, which turns every fixture into a free correctness oracle
    for our formula engine.
    """
    if not shutil.which("soffice"):
        return None
    dst = outdir / f"{src.stem}.calc.xlsx"
    try:
        _soffice(["--convert-to", "xlsx:Calc MS Excel 2007 XML",
                  "--outdir", str(outdir / "_recalc"), str(src)])
        produced = outdir / "_recalc" / f"{src.stem}.xlsx"
        if produced.exists():
            shutil.move(str(produced), str(dst))
            return dst.name
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as exc:
        print(f"  ! recalc {src.name} failed: {exc}", file=sys.stderr)
    return None


BUILDERS = {
    "basic-types": basic_types,
    "formulas": formulas,
    "styling": styling,
    "features": features,
    "edge-cases": edge_cases,
}


def main() -> int:
    OUT.mkdir(parents=True, exist_ok=True)
    (OUT / "_recalc").mkdir(exist_ok=True)
    manifest = {}
    for name, fn in BUILDERS.items():
        path = OUT / f"{name}.xlsx"
        print(f"building {path.name}")
        meta = fn(path)
        meta["bytes"] = path.stat().st_size
        meta["converted"] = convert(path, OUT)
        rc = recalc(path, OUT)
        if rc:
            meta["recalculated"] = rc
        manifest[name] = meta

    big = OUT / "large.xlsx"
    print(f"building {big.name}")
    meta = large(big)
    meta["bytes"] = big.stat().st_size
    manifest["large"] = meta

    shutil.rmtree(OUT / "_recalc", ignore_errors=True)
    (OUT / "manifest.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False))
    print(json.dumps(manifest, indent=2, ensure_ascii=False)[:2000])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
